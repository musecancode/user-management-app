from fastapi import FastAPI, Depends, UploadFile, File, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
import openpyxl
import re
from io import BytesIO
from typing import List

import models, schemas, crud
from database import engine, SessionLocal

models.Base.metadata.create_all(bind=engine)

app = FastAPI()

# ✅ CORS configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ✅ DB dependency
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

@app.get("/")
def read_root():
    return {"message": "User Management API is running"}

# ✅ Step 6: Users with search and pagination
@app.get("/users", response_model=list[schemas.UserResponse])
def read_users(
    skip: int = 0,
    limit: int = 10,
    search: str = "",
    db: Session = Depends(get_db),
):
    query = db.query(models.User)
    if search:
        query = query.filter(
            (models.User.first_name.ilike(f"%{search}%")) |
            (models.User.last_name.ilike(f"%{search}%")) |
            (models.User.email.ilike(f"%{search}%")) |
            (models.User.phone.ilike(f"%{search}%"))
        )
    return query.offset(skip).limit(limit).all()

# ✅ Create user
@app.post("/users", response_model=schemas.UserResponse)
def create_user(user: schemas.UserCreate, db: Session = Depends(get_db)):
    return crud.create_user(db, user)

# ✅ Update user
@app.put("/users/{user_id}", response_model=schemas.UserResponse)
def update_user(user_id: int, updated_user: schemas.UserCreate, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    for key, value in updated_user.dict().items():
        setattr(user, key, value)

    db.commit()
    db.refresh(user)
    return user

# ✅ Delete user
@app.delete("/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    db.delete(user)
    db.commit()
    return {"message": "User deleted successfully"}

@app.post("/upload")
def upload_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are allowed")

    try:
        contents = file.file.read()
        workbook = openpyxl.load_workbook(filename=BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail="Failed to read Excel file: " + str(e))

    sheet = workbook.active

    required_headers = ["First Name", "Last Name", "Email", "Phone Number", "PAN Number"]
    header = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]

    if [h.lower() for h in header] != [h.lower() for h in required_headers]:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid header format. Expected: {', '.join(required_headers)}"
        )

    errors = []

    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        first_name, last_name, email, phone, pan = row

        row_errors = []

        if not first_name:
            row_errors.append("Missing First Name")
        if not last_name:
            row_errors.append("Missing Last Name")
        if not email or '@' not in email:
            row_errors.append("Invalid Email")
        if not phone or not str(phone).isdigit() or len(str(phone)) != 10:
            row_errors.append("Invalid Phone Number")
        if not isinstance(pan, str) or not re.fullmatch(r"[A-Z]{5}[0-9]{4}[A-Z]", pan):
            row_errors.append("Invalid PAN format (e.g., ABCDE1234F)")

        if row_errors:
            errors.append(f"Row {i}: {', '.join(row_errors)}")

    if errors:
        raise HTTPException(status_code=400, detail=errors)

    # ✅ Insert into DB only if no errors
    for row in sheet.iter_rows(min_row=2, values_only=True):
        db_user = models.User(
            first_name=row[0],
            last_name=row[1],
            email=row[2],
            phone=str(row[3]),
            pan=row[4],
        )
        db.add(db_user)

    db.commit()
    return {"message": "Bulk upload successful"}

# ✅ Step 5.4: Download Sample Excel File
@app.get("/sample")
def download_sample_file():
    return FileResponse("sample.xlsx", filename="sample.xlsx")
